import openpyxl
#*****************************************************find search criteria*****************************************************
def get_type_criteria(search_criteria_path):
    incident_type_criteria = []
    wb = openpyxl.load_workbook(search_criteria_path)
    search_criteria_sheet = wb['Incident Types']

    for row in search_criteria_sheet.iter_rows(min_row=1, values_only=True):
        if len(row) > 1 and row[0] is not None:
            if str(row[0]).strip().upper() == 'TRUE':
                incident_type_criteria.append(row[1])

    return incident_type_criteria
#*****************************************************find region list*****************************************************
def get_region_list(search_criteria_path):
    custom_region_list = []

    wb = openpyxl.load_workbook(search_criteria_path)
    region_criteria_sheet = wb['Regions']

    for row in region_criteria_sheet.iter_rows(min_row=2, values_only=True):
        if len(row) > 1 and row[0] is not None:
            if str(row[0]).strip().upper() == 'TRUE':
                custom_region_list.append(row[2])

    return custom_region_list

