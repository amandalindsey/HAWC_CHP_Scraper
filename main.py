#*******************************************************IMPORTS*******************************************************
from file_info import incidents_path, search_criteria_path
from find_search_criteria import get_type_criteria, get_region_list
from file_info import incidents_path, search_criteria_path
from find_search_criteria import get_type_criteria, get_region_list

import subprocess
import openpyxl
import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
#****************************************************INITIALIZING VARIABLES*********************************************
#current date and time
date = datetime.datetime.now().strftime('%Y-%m-%d')
time = datetime.datetime.now().strftime('%H:%M:%S')
date_and_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
time = datetime.datetime.now().strftime('%H.%M.%S')

#Activity Log Lists
incidents_updated_list = []
incidents_added_list = []
incidents_archived_list = []

#Get Search Criteria
incident_type_criteria = get_type_criteria(search_criteria_path)
custom_region_list = get_region_list(search_criteria_path)

#Scraped Data
service = Service(executable_path='/usr/bin/chromedriver')
driver = webdriver.Chrome(service=service)
wait = WebDriverWait(driver, 10)
driver.get("https://cad.chp.ca.gov/Traffic.aspx")
#*************************************FIRST SCRAPE FOR ALL ACTIVE INCIDENT NUMBERS*************************************
all_CHP_incident_nos = []

# Wait for the dropdown to be clickable and select custom region
input_region = wait.until(EC.element_to_be_clickable((By.NAME, 'ddlSearches')))
select_custom_region = Select(input_region)
select_custom_region.select_by_value("1")

# Select all regions
select_all_regions_element = wait.until(EC.element_to_be_clickable((By.NAME, "lstCustomRegion")))
select_all_regions = Select(select_all_regions_element)

for option in select_all_regions.options:
    select_all_regions.select_by_value(option.get_attribute("value"))

# Click the 'Go' button and wait for the table to load
go_element = wait.until(EC.element_to_be_clickable((By.NAME, "btnGo")))
go_element.click()
wait.until(EC.presence_of_element_located((By.ID, "gvIncidents")))

# Turn off auto updates
find_auto_refresh_element = wait.until(EC.element_to_be_clickable((By.ID, "chkAutoRefresh")))
find_auto_refresh_element.click()

# Scrape the main table
for i in range(len(driver.find_elements(By.XPATH, "//table[@id='gvIncidents']/tbody/tr")) - 1):
    # Re-locate the rows on each iteration
    rows = driver.find_elements(By.XPATH, "//table[@id='gvIncidents']/tbody/tr")[1:]  # Skip header row
    row = rows[i]
    cells = row.find_elements(By.TAG_NAME, "td")
    incident_no = cells[1].text
    all_CHP_incident_nos.append(incident_no)

back_element = wait.until(EC.element_to_be_clickable((By.NAME, "btnEditCRList")))
back_element.click()
wait.until(EC.presence_of_element_located((By.NAME, "ddlSearches")))

#SELECT CUSTOM REGIONS FROM SEARCH CRITERIA
input_region = wait.until(EC.element_to_be_clickable((By.NAME, 'ddlSearches')))
select_custom_region = Select(input_region)
select_custom_region.select_by_value("1")
#**************************************************SELECT CUSTOM REGIONS**************************************************
select_all_regions_element = wait.until(EC.element_to_be_clickable((By.NAME, "lstCustomRegion")))
select_all_regions = Select(select_all_regions_element)

for option in select_all_regions.options:
    if option.get_attribute("value") in custom_region_list:
        select_all_regions.select_by_value(option.get_attribute("value"))

#Click GO and generate table
go_element = wait.until(EC.element_to_be_clickable((By.NAME, "btnGo")))
go_element.click()
wait.until(EC.presence_of_element_located((By.ID, "gvIncidents")))

# Turn off auto updates
#find_auto_refresh_element = wait.until(EC.element_to_be_clickable((By.ID, "chkAutoRefresh")))
#find_auto_refresh_element.click()

# Scrape the main table
scraped_data = {}

for i in range(len(driver.find_elements(By.XPATH, "//table[@id='gvIncidents']/tbody/tr")) - 1):
    rows = driver.find_elements(By.XPATH, "//table[@id='gvIncidents']/tbody/tr")[1:]  # Skip header row
    row = rows[i]
    cells = row.find_elements(By.TAG_NAME, "td")

    if len(cells) > 6 and cells[3].text in incident_type_criteria:
        incident_no = cells[1].text
        time = cells[2].text
        incident_type = cells[3].text
        location = cells[4].text
        location_desc = cells[5].text
        area = cells[6].text

        details_link = cells[0].find_element(By.TAG_NAME, "a")
        driver.execute_script("arguments[0].click();", details_link)

        #scrape lat/long
        wait.until(EC.visibility_of_element_located((By.ID, "pnlDetailsHeader")))
        detail_header = driver.find_element(By.ID, 'pnlDetailsHeader')
        wait.until(EC.visibility_of_element_located((By.ID, "lblLatLon")))
        lat_long = detail_header.find_element(By.ID, 'lblLatLon').text

        # Scrape details
        wait.until(EC.visibility_of_element_located((By.ID, "pnlDetailInfo")))

        detail_table = driver.find_element(By.ID, 'tblDetails')
        detail_tbody = detail_table.find_element(By.TAG_NAME, 'tbody')
        detail_rows = detail_tbody.find_elements(By.TAG_NAME, 'tr')
        detail_table_text = ""

        for detail_row in detail_rows:
            detail_cells = detail_row.find_elements(By.TAG_NAME, 'td')
            detail_row_text = [cell.text for cell in detail_cells]
            detail_table_text += ', '.join(detail_row_text) + '\n'

        combined_info = detail_table_text.strip()

        # Add the scraped data to scraped_data dictionary
        scraped_data[incident_no] = {
            'Incident No.': incident_no,
            'Type': incident_type,
            'Time': time,
            'Location': location,
            'Location Desc': location_desc, 'Lat/Long': lat_long,
            'Area': area,
            'Detail and Unit Information': combined_info,
            'Status': ''
        }

driver.quit()
#*************************************************OPEN EXISTING INCIDENT DATA******************************************
wb = openpyxl.load_workbook(incidents_path)

incidents_sheet = wb['Incidents']
archive_sheet = wb['Archive']
activity_log_sheet = wb['Activity Log']

headers = [cell.value for cell in incidents_sheet[1]]
#****************************************************ARCHIVE OLD INCIDENTS********************************************
incident_row_mapping = {}
rows_to_delete = []

for idx, row in enumerate(incidents_sheet.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] not in all_CHP_incident_nos:
        incidents_archived_list.append(str(incidents_sheet.cell(row=idx, column=1).value))

        archive_row_data = list(row)
        archive_row_data[8] = "Archived"
        archive_row_data.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        archive_sheet.append(archive_row_data)
        rows_to_delete.append(idx)

    incident_no = row[headers.index('Incident No.')]
    if incident_no:
        incident_row_mapping[incident_no] = idx

# Delete rows after iteration
for idx in reversed(rows_to_delete):
    incidents_sheet.delete_rows(idx)
    
#wb.save(incidents_path+str(time))
wb.save(incidents_path)
#***************************************************PROCESS NEW DATA***************************************************
#Set up Incident Data Dictionary
incident_data = {}

for row in incidents_sheet.iter_rows(min_row=2, values_only=True):
    incident_details = dict(zip(headers, row))
    incident_no = incident_details.get('Incident No.')  # or any unique identifier for each row
    if incident_no:
        incident_data[incident_no] = incident_details

#Get column indexes
detail_col = headers.index('Detail and Unit Information') + 1
status_col = headers.index('Status') + 1
date_update_col = headers.index('Last Update') + 1

#Process updates
for incident_no, incident_details in incident_data.items():
    if incident_details['Type'] not in incident_type_criteria:
        incident_details['Status'] = 'NOTE: Type excluded from search list. This incident will no longer update unless search criteria is modified.'
        row_number = incident_row_mapping.get(incident_no)
        if row_number:
            incidents_sheet.cell(row=row_number, column=status_col).value = incident_details['Status']

for incident_no, incident_details in scraped_data.items():
    if incident_no in incident_data:
        if incident_data[incident_no]['Detail and Unit Information'] != incident_details['Detail and Unit Information']:
            incident_details['Status'] = "Item Updated"
            incidents_updated_list.append(incident_no)

            row_number = incident_row_mapping.get(incident_no)
            if row_number:  # Check if the row_number exists
                # Update Detail and Unit Information
                incidents_sheet.cell(row=row_number, column=detail_col).value = incident_details['Detail and Unit Information']
                # Update Date of Last Update
                incidents_sheet.cell(row=row_number, column=date_update_col).value = date_and_time

        if incident_data[incident_no]['Detail and Unit Information'] == incident_details['Detail and Unit Information']:
            incident_details['Status'] = "No New Updates"

            row_number = incident_row_mapping.get(incident_no)
            if row_number:
                incidents_sheet.cell(row=row_number, column=status_col).value = incident_details['Status']

        # Else, if the information is the same, do nothing

    elif incident_no not in incident_data:
        incident_details['Status'] = "New Item"
        incidents_added_list.append(incident_no)
        incident_details['Last Update'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            # Convert dict_values to a list before appending
            incidents_sheet.append(list(incident_details.values()))
        except Exception as e:
            print(f"Error appending new data for {incident_no}: {e}")

    else:
        # Handle any other unexpected cases
        print(f"Unexpected case encountered for incident number: {incident_no}")

#wb.save(incidents_path+str(time))
wb.save(incidents_path)
#**************************************************UPDATE ACTIVITY LOG**************************************************
#activity_headers = [cell.value for cell in activity_log_sheet[1]]

updated_list = '; '.join(incidents_updated_list)
if updated_list == "":
    updated_list = "null"

archived_list = '; '.join(incidents_archived_list)
if archived_list == "":
    archived_list = "null"

added_list = '; '.join(incidents_added_list)
if added_list == "":
    added_list = "null"

activity_log_sheet.append([date, time, updated_list, added_list, archived_list])

#wb.save(incidents_path+str(time))
wb.save(incidents_path)
wb.close()

#subprocess.run(["open", incidents_path+str(time)])




